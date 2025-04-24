"""
Subscriber service module.
"""

from dataclasses import dataclass
from datetime import datetime
from enum import Enum
from io import BytesIO
from typing import Any, Dict, Tuple

from marshmallow import ValidationError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from psycopg2.errors import UniqueViolation
from sqlalchemy.exc import IntegrityError

from app.models import Subscriber
from app.queue import QueueService
from app.schemas import SubscriberSchema
from config import Config
from utils.db_utils import paginate_query
from utils.email_validation import validate_email
from utils.service_base import BaseService


class EmailTemplate(Enum):
    """Email template types"""

    CONFIRMATION = "email_confirmation_email_form.html"
    PASSWORD_RESET = "password_reset_email_form.html"  # nosec
    CONTACT_US = "contact_us_email-form.html"
    WELCOME = "welcome_email-form.html"
    BROADCAST = "broadcast_email-form.html"


@dataclass
class EmailConfig:
    """Email configuration data structure"""

    service: str
    recipient: str
    subject: str
    template: str
    data: Dict[str, str]


class SubscriberService(BaseService):
    """Subscriber service class."""

    def __init__(self, page_size: int = 10):
        """Initialize subscriber service."""
        super().__init__(Subscriber, SubscriberSchema, page_size)
        self.queue_service = QueueService()
        self.sender_email = Config.SMTP_USER

    def create_subscriber(self, form_data: Dict[str, Any]) -> Subscriber:
        """
        Create a new subscriber.

        Args:
            form_data: Dictionary containing subscriber data from the form.

        Returns:
            The created Subscriber instance.

        Raises:
            ValidationError: If validation fails.
        """
        try:
            # Process and validate form data
            processed_data = self.validate_form_data(form_data)

            # Validate with schema
            self.validate_with_schema(processed_data)

            # Validate email
            validate_email(processed_data["email"])

            # Create the subscriber
            subscriber = self.model_class()
            try:
                subscriber.create(**processed_data)
            except IntegrityError as db_error:
                if isinstance(db_error.orig, UniqueViolation):
                    raise UniqueViolation(
                        "Email already subscribed"
                    ) from db_error
                raise db_error

            # Send a welcome email
            email_config = EmailConfig(
                service="SWEA",
                recipient=processed_data["email"],
                subject="Welcome to SWEA!",
                template=EmailTemplate.WELCOME.value,
                data={
                    "recipient": processed_data["email"].split("@")[0],
                    "unsubscribe_link": f"{Config.BASE_URL}/unsubscribe?email={processed_data['email']}",
                },
            )

            self.queue_service.enqueue_task(
                "send_email", email_config.__dict__
            )

            return subscriber

        except ValidationError as error:
            raise ValidationError(error.messages) from error
        except ValueError as ve:
            raise ValueError(ve) from ve

    def search_subscribers_by_email(self, email: str) -> Dict[str, Any]:
        """
        Search for a subscriber by email.

        Args:
            email: The email address to search for.

        Returns:
            The Subscriber instance if found, None otherwise.
        """
        return paginate_query(self.model_class, email=email)

    def delete_subscriber(self, subscriber_email: str) -> bool:
        """
        Delete a subscriber.

        Args:
            subscriber_email: The email of the subscriber to delete.

        Returns:
            True if subscriber was deleted, False if not found.
        """
        try:
            validate_email(subscriber_email)
            subscriber = self.search_subscribers_by_email(subscriber_email)

            if subscriber["total_items"] > 0:
                uuid = subscriber["data"][0]["uuid"]
                self.delete(uuid, permanent=True)
                return True

            return False

        except Exception as e:
            raise Exception(e) from e

    def send_broadcast_email(self, form_data: Dict[str, Any]) -> None:
        """
        Send a broadcast email to all subscribers.
        Includes input validation and error handling.
        """
        try:
            email_data = self.validate_form_data(
                form_data, validate="broadcast"
            )

            # Get all subscribers
            subscribers = self.get_all()

            if subscribers["total_items"] == 0:
                raise ValueError("No subscribers found")

            list_of_subscribers = [
                subscriber["email"] for subscriber in subscribers["data"]
            ]

            for subscriber in list_of_subscribers:
                email_config = EmailConfig(
                    service="SWEA",
                    recipient=subscriber,
                    subject=email_data["subject"],
                    template=EmailTemplate.BROADCAST.value,
                    data={
                        "subject": email_data["subject"],
                        "recipient": subscriber.split("@")[0],
                        "message": email_data["message"],
                        "unsubscribe_link": f"{Config.BASE_URL}/unsubscribe?email={subscriber}",
                    },
                )

                self.queue_service.enqueue_task(
                    "send_email", email_config.__dict__
                )

        except ValueError as ve:
            raise ValueError(f"Invalid broadcast email: {ve}") from ve
        except Exception as e:
            raise Exception(f"Failed to send broadcast email: {e}") from e

    def validate_form_data(
        self, form_data: Dict[str, Any], validate: str = "email"
    ) -> Dict[str, Any]:
        """
        Validate the form data submitted by the user.
        Raises ValueError if validation fails.
        """
        if validate == "email":
            required_keys = {"email"}
            missing_keys = required_keys - form_data.keys()

            if missing_keys:
                raise ValueError(
                    f"Missing required fields: {', '.join(missing_keys)}"
                )

            return form_data

        required_keys = {"subject", "message"}
        missing_keys = required_keys - form_data.keys()

        if missing_keys:
            raise ValueError(
                f"Missing required fields: {', '.join(missing_keys)}"
            )

        return form_data

    def export_subscribers(self) -> Tuple[BytesIO, str]:
        """
        Export all subscribers to a professionally formatted Excel file.

        Returns:
            A tuple containing the BytesIO object and the filename.

        """
        subscribers = self.get_all().get("data", [])

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "SWEA Subscribers"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")  # White text for headers
        header_fill = PatternFill(
            start_color="ff520000", end_color="ff520000", fill_type="solid"
        )  # Dark red fill for headers
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add headers
        headers = ["Email", "Created At"]
        ws.append(headers)

        # Apply header styles
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border

        # Add subscriber data
        for subscriber in subscribers:
            ws.append(
                [
                    subscriber.get("email"),
                    (
                        subscriber.get("created_at").strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                        if subscriber.get("created_at")
                        else ""
                    ),
                ]
            )

        # Apply styles to data rows
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center"
                )
                cell.border = thin_border

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    print(e)
                    max_length = 0

            adjusted_width = (max_length + 2) * 1.2  # Add some padding
            ws.column_dimensions[column].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Save the workbook to a BytesIO object
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        date_str = datetime.now().strftime("%Y-%m-%d")

        return output, f"SWEA_Subscribers_{date_str}.xlsx"
